import shutil
from flask import Flask, request, jsonify, send_file, render_template
import openpyxl
from openpyxl.styles import NamedStyle
import os
from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill, Border
from openpyxl.drawing.image import Image

app = Flask(__name__)

# Path template Excel dan folder output
TEMPLATE_PATH = "templates/laporan_template.xlsx"
OUTPUT_DIR = "output"

if not os.path.exists(OUTPUT_DIR):
    os.makedirs(OUTPUT_DIR)

def copy_template_to_sheet(template_ws, target_ws):
    """Salin semua konten, gaya, pengaturan halaman, dimensi, merge cells, dan gambar dari template_ws ke target_ws."""
    for row in template_ws.iter_rows():
        for cell in row:
            new_cell = target_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            
            # Salin Font
            if cell.font:
                new_cell.font = Font(
                    name=cell.font.name,
                    size=cell.font.size,
                    bold=cell.font.bold,
                    italic=cell.font.italic,
                    underline=cell.font.underline,
                    strike=cell.font.strike,
                    color=cell.font.color
                )

            # Salin Alignment
            if cell.alignment:
                new_cell.alignment = Alignment(
                    horizontal=cell.alignment.horizontal,
                    vertical=cell.alignment.vertical,
                    wrap_text=cell.alignment.wrap_text,
                    shrink_to_fit=cell.alignment.shrink_to_fit,
                    text_rotation=cell.alignment.text_rotation
                )

            # Salin Fill
            # Salin Fill


            # Salin Border
            if cell.border:
                sides = ['left', 'right', 'top', 'bottom', 'diagonal']
                border_sides = {side: getattr(cell.border, side) for side in sides}
                new_cell.border = Border(
                    left=border_sides['left'],
                    right=border_sides['right'],
                    top=border_sides['top'],
                    bottom=border_sides['bottom'],
                    diagonal=border_sides['diagonal'],
                    diagonal_direction=cell.border.diagonal_direction
                )

    # Salin merge cells
    for merged_range in template_ws.merged_cells.ranges:
        target_ws.merge_cells(str(merged_range))

    # Salin pengaturan halaman dan margin
    target_ws.page_setup = template_ws.page_setup
    target_ws.print_options = template_ws.print_options
    target_ws.sheet_properties = template_ws.sheet_properties
    target_ws.page_margins = template_ws.page_margins

    # Atur Header dan Footer secara manual (jika diperlukan)
    if hasattr(template_ws, 'header_footer'):
        target_ws.header_footer = template_ws.header_footer

    # Salin dimensi kolom dari template ke target
    for col_letter, col_dimension in template_ws.column_dimensions.items():
        if col_dimension.width:
            target_ws.column_dimensions[col_letter].width = col_dimension.width

    # Salin dimensi baris
    for row_number, row_dimension in template_ws.row_dimensions.items():
        if row_dimension.height:
            target_ws.row_dimensions[row_number].height = row_dimension.height

    # Salin gambar
    for image in template_ws._images:
        img = Image(image.ref)
        img.anchor = image.anchor
        target_ws.add_image(img)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/input', methods=['POST'])
def input_data():
    try:
        action = request.form.get('action')
        proyek = request.form.get('proyek', '')
        customer = request.form.get('customer', '')
        tempat_test = request.form.get('tempat_test', '')
        tanggal_test = request.form.get('tanggal_test', '')

        kode_benda_uji = request.form.getlist('kode_benda_uji')
        mutu = request.form.getlist('mutu')
        umur_test = request.form.getlist('umur_test')
        tanggal_cor = request.form.getlist('tanggal_cor')
        jenis_benda_uji = request.form.getlist('jenis_benda_uji')
        ukuran_benda_uji = request.form.getlist('ukuran_benda_uji')
        volume = [float(v) if v else 0 for v in request.form.getlist('volume')]
        berat = [float(b) if b else 0 for b in request.form.getlist('berat')]
        beban = [float(b) if b else 0 for b in request.form.getlist('beban')]
        tipe_retakan = request.form.getlist('tipe_retakan')

        row_count = len(kode_benda_uji)

        tanggal_test_formatted = datetime.strptime(tanggal_test, '%Y-%m-%d').strftime('%d-%m-%Y')
        bulan_ini = datetime.strptime(tanggal_test, '%Y-%m-%d').strftime('%Y-%m')

        output_path = os.path.join(OUTPUT_DIR, f"laporan_{bulan_ini}.xlsx")

        if action == 'new_file':
            if os.path.exists(output_path):
                os.remove(output_path)
            wb = openpyxl.load_workbook(TEMPLATE_PATH)
            ws = wb.active
            ws.title = tanggal_test_formatted
        elif action == 'new_sheet':
            if os.path.exists(output_path):
                wb = openpyxl.load_workbook(output_path)
                if tanggal_test_formatted in wb.sheetnames:
                    return jsonify({"error": f"Sheet {tanggal_test_formatted} sudah ada."}), 400
                ws = wb.create_sheet(title=tanggal_test_formatted)
            else:
                return jsonify({"error": "File tidak ditemukan. Pilih 'Buat file baru'."}), 400
       
        wb.save(output_path)
        return f"<script>alert('Data berhasil dimasukkan!'); window.location.href='/'</script>"
    except Exception as e:
        return f"<script>alert('Terjadi kesalahan: {str(e)}'); window.location.href='/'</script>"
@app.route('/download/<filename>')
def download_file(filename):
    file_path = os.path.join(OUTPUT_DIR, filename)
    if os.path.exists(file_path):
        return send_file(file_path, as_attachment=True)
    return jsonify({"error": "File tidak ditemukan"}), 404

if __name__ == '__main__':
    app.run(debug=True)
